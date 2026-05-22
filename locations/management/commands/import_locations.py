# -*- coding: utf-8 -*-
"""
Management command to bulk-import Locations from an Excel file.

Expected columns (matching the Locations_Bulk_Upload.xlsx format):
  Title, Description, Keywords, Category, Location, Reference,
  Labels, Image, Video, Visible, Private, Feed URL, Rate, Level,
  Is Verified, Data Source

Usage:
    python manage.py import_locations /path/to/Locations_Bulk_Upload.xlsx
"""
import hashlib

import openpyxl

from django.core.management.base import BaseCommand, CommandError

from locations.models import Location


# Maps Excel column header → model field name.
COLUMN_MAP = {
    'Title': 'name',
    'Description': 'description',
    'Keywords': 'keywords',
    'Category': 'object_type',
    'Reference': 'abbreviation',
    'Labels': 'labels',
    'Image': 'image',
    'Video': 'video',
    'Feed URL': 'feed_url',
    'Rate': 'rate',
    'Level': 'level',
    'Data Source': 'data_source',
}

BOOLEAN_COLUMNS = {'Visible', 'Private', 'Is Verified'}


def _parse_bool(value):
    if isinstance(value, bool):
        return value
    if isinstance(value, str):
        return value.strip().lower() in ('true', '1', 'yes')
    if isinstance(value, int):
        return bool(value)
    return False


def _compute_import_key(name, category, location):
    """
    Return an MD5 hex digest for the imported row.

    For valid coordinates, key on a normalized "lat,lng" string so the
    Excel importer can continue to match records previously imported
    from the map API. For missing or malformed coordinates, fall back
    to additional stable fields so blank locations do not all collapse
    to the same import key.
    """
    lat, lng = _parse_coords(location)
    if lat is not None and lng is not None:
        raw = 'coords:{0:.15g},{1:.15g}'.format(lat, lng)
    else:
        raw = 'fallback:{0}|{1}|{2}'.format(
            str(name or '').strip(),
            str(category or '').strip(),
            str(location or '').strip(),
        )
    return hashlib.md5(raw.encode('utf-8')).hexdigest()


def _parse_coords(value):
    """
    Parse a "lat,lng" string into a (lat, lng) tuple of Decimals,
    or (None, None) if the value is absent or malformed.
    """
    if not value:
        return None, None
    try:
        parts = str(value).split(',')
        if len(parts) == 2:
            return float(parts[0].strip()), float(parts[1].strip())
    except (ValueError, AttributeError):
        pass
    return None, None


class Command(BaseCommand):
    help = 'Import locations from a bulk-upload Excel (.xlsx) file.'

    def add_arguments(self, parser):
        parser.add_argument(
            'file',
            type=str,
            help='Path to the Excel file to import.'
        )
        parser.add_argument(
            '--dry-run',
            action='store_true',
            dest='dry_run',
            default=False,
            help='Parse the file and report what would be imported without saving.'
        )

    def handle(self, *args, **options):
        filepath = options['file']
        dry_run = options['dry_run']

        try:
            wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        except Exception as e:
            raise CommandError('Could not open Excel file: {}'.format(e))

        ws = wb.active

        # Build header → column-index mapping from the first row.
        headers = []
        for cell in next(ws.iter_rows(min_row=1, max_row=1)):
            headers.append(cell.value)

        created = 0
        updated = 0
        skipped = 0
        deleted = 0
        seen_keys = set()

        for row in ws.iter_rows(min_row=2, values_only=True):
            row_data = dict(zip(headers, row))

            name = row_data.get('Title')
            if not name:
                skipped += 1
                continue

            name = str(name).strip()
            if not name:
                skipped += 1
                continue

            defaults = {}

            for col_header, field_name in COLUMN_MAP.items():
                val = row_data.get(col_header)
                if val is not None:
                    defaults[field_name] = str(val).strip() if not isinstance(val, bool) else val

            # Boolean fields
            defaults['visible'] = _parse_bool(row_data.get('Visible', True))
            defaults['private'] = _parse_bool(row_data.get('Private', False))
            defaults['is_verified'] = _parse_bool(row_data.get('Is Verified', False))

            # Location column: "lat,lng"
            lat, lng = _parse_coords(row_data.get('Location'))
            if lat is not None:
                defaults['googlemap_lat'] = lat
                defaults['googlemap_lng'] = lng

            import_key = _compute_import_key(
                name,
                row_data.get('Category'),
                row_data.get('Location'),
            )
            defaults['import_key'] = import_key
            seen_keys.add(import_key)

            if dry_run:
                self.stdout.write('Would import: {} ({})'.format(
                    name, defaults.get('object_type', 'no category')
                ))
                continue

            obj, was_created = Location.objects.update_or_create(
                import_key=import_key,
                defaults=defaults
            )

            if was_created:
                created += 1
            else:
                updated += 1

        wb.close()

        # Remove Excel-imported records that were not present in this file.
        # Records without an import_key (e.g. from the map API) are left untouched.
        stale_qs = Location.objects.filter(
            import_key__isnull=False
        ).exclude(import_key__in=seen_keys)

        if dry_run:
            stale_names = list(stale_qs.values_list('name', flat=True))
            for stale_name in stale_names:
                self.stdout.write('Would delete: {}'.format(stale_name))
            self.stdout.write(self.style.WARNING(
                'Dry run complete. No records were saved. {} would be deleted.'.format(
                    len(stale_names)
                )
            ))
        else:
            deleted, _ = stale_qs.delete()
            self.stdout.write(self.style.SUCCESS(
                'Import complete. Created: {}, Updated: {}, Deleted: {}, Skipped: {}'.format(
                    created, updated, deleted, skipped
                )
            ))
